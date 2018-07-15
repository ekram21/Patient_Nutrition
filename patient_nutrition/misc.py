
'''a library of functions to carry out tasks on the table'''

import sqlite3
from openpyxl import load_workbook

def Get_All_TableNames(database_name):
	output_array = []
	con = sqlite3.connect(database_name + '.db')
	cursor = con.cursor()
	cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
	lol = (cursor.fetchall())

	for line in lol:
		for l in line:
			output_array.append(l)

	return output_array

def Get_Table_Headings_From_Table(database_name, table_name):
	print ('Reading from database: ' + database_name + ' and table: ' + table_name)

	print('Returning this tables headings in an array format')

	conn = sqlite3.connect( database_name + '.db')

	c = conn.cursor()

	# Select everything from table requested

	c.execute("SELECT * FROM " + table_name)

	#Format and print column names..
	table_headings = [i[0] for i in c.description]

	return table_headings

def Get_Data_From_Table(database_name, table_name):
	print ('Reading from database: ' + database_name + ' and table: ' + table_name)

	print('Returning this tables data in an array format')

	conn = sqlite3.connect( database_name + '.db')

	c = conn.cursor()

	# Select everything from table requested
	c.execute("SELECT * FROM " + table_name)

	rows = c.fetchall()

	Table_Data = []

	for l in rows:
		Table_Data.append(list(l))

	return Table_Data

def Insert_RowOFData_in_Table(database_name, table_name, row_data_Array):
	print ('Reading from database: ' + database_name + ' and table: ' + table_name)

	print('Inserting accepted data array into specified table..')

	conn = sqlite3.connect( database_name + '.db')

	c = conn.cursor()

	#Craft comma separated string from accepted row data array
	crafted_String = ''
	for elem in row_data_Array:
		if str(elem).isdigit():
			#the elem in question contains only digits
			crafted_String = crafted_String + str(elem) + ','
		else:
			#the elem in question contains word characters
			crafted_String = crafted_String + '"' + str(elem) + '"' + ','

	crafted_String = crafted_String[0:-1]

	print (crafted_String)

	c.execute("INSERT INTO " + str(table_name) + " VALUES (" + crafted_String +  ")")

	# Save (commit) the changes
	conn.commit()

	# We can also close the connection if we are done with it.
	conn.close()

def Clear_Specified_Table(database_name, table_name):
	print ('Reading from database: ' + database_name + ' and table: ' + table_name)

	print('Inserting accepted data array into specified table..')

	conn = sqlite3.connect( database_name + '.db')

	c = conn.cursor()

	myQuery = 'DELETE FROM ' + str(table_name)

	print (myQuery)

	c.execute(myQuery)

	# Save (commit) the changes
	conn.commit()

	# We can also close the connection if we are done with it.
	conn.close()



def Get_All_Data_From_Excel_File(fileName, SheetName):
	dataSet = []

	# Load in the workbook
	wb = load_workbook('./' + fileName)

	sheet = wb[SheetName]

	rowNum = sheet.max_row
	columnNum = sheet.max_column

	for z in range(1, rowNum+1):
		this_array = []
		for i in range(1, columnNum+1):
			curr = sheet.cell(row=z, column=i).value
			this_array.append(curr)

		dataSet.append(this_array)

	return dataSet



if __name__ == '__main__':
	lol = Get_All_TableNames('FixedDatabase')
	print (lol)



